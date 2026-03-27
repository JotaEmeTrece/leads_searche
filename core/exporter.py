"""
Exportador de leads desde SQLite a Excel.
Mantiene el formato actual del reporte.
"""

from datetime import datetime
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.storage import get_all_leads


def _filtrar_leads(leads: List[Dict[str, str]], ciudad: Optional[str] = None, fuente: Optional[str] = None) -> List[Dict[str, str]]:
    filtrados = leads
    if fuente:
        filtrados = [lead for lead in filtrados if str(lead.get("fuente", "")).lower() == fuente.lower()]
    if ciudad:
        filtrados = [lead for lead in filtrados if str(lead.get("ciudad", "")).lower() == ciudad.lower()]
    return filtrados


def exportar_excel(leads: List[Dict[str, str]], archivo: str, tipo_negocio: str, ciudad: str) -> None:
    """Exporta leads a Excel con el mismo formato visual del script original."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    verde_claro = PatternFill("solid", fgColor="e8f5e9")
    gris_header = PatternFill("solid", fgColor="1a1f2e")
    amarillo = PatternFill("solid", fgColor="fff3cd")
    rojo_claro = PatternFill("solid", fgColor="fdecea")
    blanco = PatternFill("solid", fgColor="ffffff")
    gris_fila = PatternFill("solid", fgColor="f8f9fa")

    borde = Border(
        left=Side(style="thin", color="dee2e6"),
        right=Side(style="thin", color="dee2e6"),
        top=Side(style="thin", color="dee2e6"),
        bottom=Side(style="thin", color="dee2e6"),
    )

    ws.merge_cells("A1:J1")
    ws["A1"] = f"LEADS - {tipo_negocio.upper()} EN {ciudad.upper()} - {datetime.now().strftime('%d/%m/%Y')}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = gris_header
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    total = len(leads)
    con_wa = sum(1 for lead in leads if lead.get("whatsapp") == "SI")
    probable = sum(1 for lead in leads if lead.get("whatsapp") == "PROBABLE")
    ws["A2"] = (
        f"Total encontrados: {total}  |  Con WhatsApp confirmado: {con_wa}  |  "
        f"WhatsApp probable: {probable}  |  Generado por Jercol Technologies"
    )
    ws["A2"].font = Font(size=9, color="5a6478")
    ws["A2"].fill = PatternFill("solid", fgColor="f4f6f9")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    headers = ["#", "Negocio", "Categoria", "Telefono", "WhatsApp", "Direccion", "Website", "Rating", "Resenas", "Notas"]
    ws.append([])
    ws.append(headers)
    header_row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = gris_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde
    ws.row_dimensions[header_row].height = 22

    for i, lead in enumerate(leads, 1):
        row_num = header_row + i
        wa = lead.get("whatsapp", "NO DETECTADO")
        telefono = lead.get("telefono") if lead.get("telefono") else "N/A"

        row = [
            i,
            lead.get("nombre", "N/A"),
            lead.get("categoria", "N/A"),
            telefono,
            wa,
            lead.get("direccion", "N/A"),
            lead.get("website", "N/A"),
            lead.get("rating", "N/A"),
            lead.get("resenas", "N/A"),
            lead.get("notas", ""),
        ]
        ws.append(row)

        fill_row = gris_fila if i % 2 == 0 else blanco
        for col in range(1, 11):
            cell = ws.cell(row=row_num, column=col)
            cell.border = borde
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(size=9.5)
            if col != 5:
                cell.fill = fill_row

        wa_cell = ws.cell(row=row_num, column=5)
        wa_cell.font = Font(bold=True, size=9.5)
        wa_cell.alignment = Alignment(horizontal="center", vertical="center")
        if wa == "SI":
            wa_cell.fill = verde_claro
            wa_cell.font = Font(bold=True, size=9.5, color="1b5e20")
        elif wa == "PROBABLE":
            wa_cell.fill = amarillo
            wa_cell.font = Font(bold=True, size=9.5, color="856404")
        else:
            wa_cell.fill = rojo_claro
            wa_cell.font = Font(bold=True, size=9.5, color="721c24")

        ws.row_dimensions[row_num].height = 20

    anchos = [5, 30, 18, 16, 14, 35, 35, 8, 10, 20]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    ws2 = wb.create_sheet("Solo WhatsApp")
    ws2.append(["#", "Negocio", "Telefono", "WhatsApp", "Direccion", "Website"])
    for col in range(1, 7):
        cell = ws2.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = gris_header
        cell.alignment = Alignment(horizontal="center")

    contador = 1
    for lead in leads:
        if lead.get("whatsapp") in ("SI", "PROBABLE"):
            ws2.append(
                [
                    contador,
                    lead.get("nombre", ""),
                    lead.get("telefono") if lead.get("telefono") else "N/A",
                    lead.get("whatsapp", ""),
                    lead.get("direccion", ""),
                    lead.get("website", ""),
                ]
            )
            contador += 1

    for col, ancho in enumerate([5, 32, 16, 14, 38, 38], 1):
        ws2.column_dimensions[get_column_letter(col)].width = ancho

    wb.save(archivo)
    print(f"\n  Excel guardado: {archivo}")
    print(f"  Total leads: {total}")
    print(f"  Con WhatsApp confirmado: {con_wa}")
    print(f"  WhatsApp probable: {probable}")


def exportar_desde_db(
    archivo: str,
    tipo_negocio: str,
    ciudad: str,
    db_path: str = "leads.db",
    fuente: Optional[str] = "maps",
) -> None:
    """Lee leads desde SQLite y exporta a Excel."""
    leads = get_all_leads(db_path=db_path)
    leads_filtrados = _filtrar_leads(leads, ciudad=ciudad, fuente=fuente)
    exportar_excel(leads_filtrados, archivo=archivo, tipo_negocio=tipo_negocio, ciudad=ciudad)
