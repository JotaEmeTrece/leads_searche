"""
JERCOL TECHNOLOGIES — Buscador de Leads con WhatsApp
=====================================================
Busca negocios en Google Maps por ciudad y tipo,
detecta si tienen WhatsApp y exporta a Excel.

USO:
    python buscador_leads.py

CONFIGURACION:
    Edita las variables en la seccion CONFIG mas abajo.
"""

from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time
import random
import re
import json
from datetime import datetime

# ─────────────────────────────────────────────────────────────
# CONFIG — edita esto antes de correr
# ─────────────────────────────────────────────────────────────
TIPO_NEGOCIO = "restaurantes"       # restaurantes, ferreterias, dentistas, peluquerias, etc.
CIUDAD       = "Monteria Colombia"  # ciudad donde buscar
MAX_RESULTS  = 10                   # cuantos negocios quieres extraer (max recomendado: 60)
ARCHIVO      = "leads_jercol_rest.xlsx"  # nombre del archivo Excel de salida

# Delays en segundos — NO reducir, son para evitar baneo
DELAY_ENTRE_CLICKS  = (2.0, 4.0)   # pausa entre clicks en resultados
DELAY_ENTRE_SCROLL  = (1.5, 3.0)   # pausa entre scrolls
DELAY_VISITA_WEB    = (2.0, 4.0)   # pausa al visitar la web del negocio
# ─────────────────────────────────────────────────────────────

def delay(rango=None):
    """Pausa aleatoria para simular comportamiento humano."""
    if rango is None:
        rango = (1.0, 2.5)
    time.sleep(random.uniform(*rango))

def es_celular_colombiano(telefono):
    """Detecta si un numero es celular colombiano (empieza con 3)."""
    if not telefono:
        return False
    limpio = re.sub(r'[\s\-\(\)\+]', '', telefono)
    # Formato colombiano: 57 3XX o 3XX
    return bool(re.match(r'^(57)?3\d{9}$', limpio))

def buscar_whatsapp_en_web(page, url):
    """
    Visita la pagina web del negocio y busca señales de WhatsApp.
    Retorna: 'si', 'probable', 'no', o 'sin_web'
    """
    if not url or url == 'N/A':
        return 'sin_web'
    try:
        delay(DELAY_VISITA_WEB)
        page.goto(url, wait_until='domcontentloaded', timeout=15000)
        delay((1.0, 2.0))
        html = page.content()
        soup = BeautifulSoup(html, 'html.parser')
        texto = html.lower()

        # Señal directa — link wa.me
        if 'wa.me' in texto or 'api.whatsapp.com' in texto:
            return 'si'

        # Señal directa — texto whatsapp cerca de un numero
        if 'whatsapp' in texto:
            return 'si'

        # Señal directa — icono de whatsapp
        if 'whatsapp' in str(soup.find_all('img')).lower():
            return 'si'

        return 'no'

    except Exception:
        return 'error'

def extraer_detalle_negocio(page):
    """
    Extrae los datos del panel lateral de un negocio en Google Maps.
    """
    delay((1.5, 3.0))

    datos = {
        'nombre':    'N/A',
        'categoria': 'N/A',
        'direccion': 'N/A',
        'telefono':  'N/A',
        'website':   'N/A',
        'rating':    'N/A',
        'resenas':   'N/A',
        'horario':   'N/A',
    }

    try:
        html = page.content()
        soup = BeautifulSoup(html, 'html.parser')

        # Nombre
        try:
            nombre = page.locator('h1').first.inner_text(timeout=5000)
            datos['nombre'] = nombre.strip()
        except Exception:
            pass

        # Rating y resenas
        try:
            rating_el = page.locator('[aria-label*="estrellas"]').first
            rating_text = rating_el.get_attribute('aria-label', timeout=3000)
            if rating_text:
                match = re.search(r'([\d,.]+)\s+estrellas.*?([\d,.]+)\s+rese', rating_text)
                if match:
                    datos['rating']  = match.group(1)
                    datos['resenas'] = match.group(2)
        except Exception:
            pass

        # Telefono — buscar en botones de llamada
        try:
            tel_patterns = [
                'button[data-tooltip*="Copiar numero"]',
                'button[aria-label*="telefono"]',
                'button[aria-label*="Telefono"]',
                '[data-item-id*="phone"]',
            ]
            for pattern in tel_patterns:
                try:
                    el = page.locator(pattern).first
                    label = el.get_attribute('aria-label', timeout=2000)
                    if label:
                        tel = re.search(r'[\+\d][\d\s\-\(\)]{7,}', label)
                        if tel:
                            datos['telefono'] = tel.group().strip()
                            break
                except Exception:
                    continue

            # Si no encontro, buscar en el HTML
            if datos['telefono'] == 'N/A':
                tel_match = re.search(r'\+?57[\s\-]?[3][0-9]{9}', html)
                if tel_match:
                    datos['telefono'] = tel_match.group().strip()
                else:
                    tel_match2 = re.search(r'3[0-9]{9}', html)
                    if tel_match2:
                        datos['telefono'] = tel_match2.group().strip()
        except Exception:
            pass

        # Direccion
        try:
            dir_patterns = [
                '[data-item-id*="address"]',
                'button[aria-label*="direcci"]',
                '[data-tooltip*="direcci"]',
            ]
            for pattern in dir_patterns:
                try:
                    el = page.locator(pattern).first
                    label = el.get_attribute('aria-label', timeout=2000)
                    if label and len(label) > 5:
                        datos['direccion'] = label.replace('Direcci\u00f3n: ', '').strip()
                        break
                except Exception:
                    continue
        except Exception:
            pass

        # Website
        try:
            web_patterns = [
                'a[data-item-id*="authority"]',
                'a[aria-label*="sitio web"]',
                'a[aria-label*="Sitio web"]',
            ]
            for pattern in web_patterns:
                try:
                    el = page.locator(pattern).first
                    href = el.get_attribute('href', timeout=2000)
                    if href and href.startswith('http'):
                        datos['website'] = href
                        break
                except Exception:
                    continue
        except Exception:
            pass

        # Categoria
        try:
            cat_el = page.locator('button[jsaction*="category"]').first
            datos['categoria'] = cat_el.inner_text(timeout=3000).strip()
        except Exception:
            pass

    except Exception as e:
        print(f"    Error extrayendo detalle: {e}")

    return datos

def determinar_whatsapp(telefono, whatsapp_web):
    """Determina el estado de WhatsApp basado en telefono y busqueda web."""
    if whatsapp_web == 'si':
        return 'SI'
    if es_celular_colombiano(telefono):
        if whatsapp_web == 'no':
            return 'PROBABLE'
        return 'PROBABLE'
    if whatsapp_web == 'probable':
        return 'PROBABLE'
    return 'NO DETECTADO'

def exportar_excel(leads, archivo):
    """Exporta los leads a un Excel con formato profesional."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Colores
    verde_oscuro  = PatternFill("solid", fgColor="00b37e")
    verde_claro   = PatternFill("solid", fgColor="e8f5e9")
    gris_header   = PatternFill("solid", fgColor="1a1f2e")
    amarillo      = PatternFill("solid", fgColor="fff3cd")
    rojo_claro    = PatternFill("solid", fgColor="fdecea")
    blanco        = PatternFill("solid", fgColor="ffffff")
    gris_fila     = PatternFill("solid", fgColor="f8f9fa")

    borde = Border(
        left=Side(style='thin', color='dee2e6'),
        right=Side(style='thin', color='dee2e6'),
        top=Side(style='thin', color='dee2e6'),
        bottom=Side(style='thin', color='dee2e6'),
    )

    # Titulo
    ws.merge_cells('A1:J1')
    ws['A1'] = f'LEADS — {TIPO_NEGOCIO.upper()} EN {CIUDAD.upper()} — {datetime.now().strftime("%d/%m/%Y")}'
    ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill = gris_header
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Subtitulo stats
    ws.merge_cells('A2:J2')
    total = len(leads)
    con_wa = sum(1 for l in leads if l.get('whatsapp') == 'SI')
    probable = sum(1 for l in leads if l.get('whatsapp') == 'PROBABLE')
    ws['A2'] = f'Total encontrados: {total}  |  Con WhatsApp confirmado: {con_wa}  |  WhatsApp probable: {probable}  |  Generado por Jercol Technologies'
    ws['A2'].font = Font(size=9, color='5a6478')
    ws['A2'].fill = PatternFill("solid", fgColor="f4f6f9")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    # Headers
    headers = ['#', 'Negocio', 'Categoria', 'Telefono', 'WhatsApp', 'Direccion', 'Website', 'Rating', 'Resenas', 'Notas']
    ws.append([])
    ws.append(headers)
    header_row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = Font(bold=True, size=10, color='FFFFFF')
        cell.fill = gris_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = borde
    ws.row_dimensions[header_row].height = 22

    # Datos
    for i, lead in enumerate(leads, 1):
        row_num = header_row + i
        wa = lead.get('whatsapp', 'NO DETECTADO')

        row = [
            i,
            lead.get('nombre', 'N/A'),
            lead.get('categoria', 'N/A'),
            lead.get('telefono', 'N/A'),
            wa,
            lead.get('direccion', 'N/A'),
            lead.get('website', 'N/A'),
            lead.get('rating', 'N/A'),
            lead.get('resenas', 'N/A'),
            lead.get('notas', ''),
        ]
        ws.append(row)

        # Color por fila y estado WA
        fill_row = gris_fila if i % 2 == 0 else blanco
        for col in range(1, 11):
            cell = ws.cell(row=row_num, column=col)
            cell.border = borde
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.font = Font(size=9.5)
            if col != 5:
                cell.fill = fill_row

        # Color celda WhatsApp
        wa_cell = ws.cell(row=row_num, column=5)
        wa_cell.font = Font(bold=True, size=9.5)
        wa_cell.alignment = Alignment(horizontal='center', vertical='center')
        if wa == 'SI':
            wa_cell.fill = verde_claro
            wa_cell.font = Font(bold=True, size=9.5, color='1b5e20')
        elif wa == 'PROBABLE':
            wa_cell.fill = amarillo
            wa_cell.font = Font(bold=True, size=9.5, color='856404')
        else:
            wa_cell.fill = rojo_claro
            wa_cell.font = Font(bold=True, size=9.5, color='721c24')

        ws.row_dimensions[row_num].height = 20

    # Anchos de columna
    anchos = [5, 30, 18, 16, 14, 35, 35, 8, 10, 20]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    # Hoja 2 — solo los que tienen WhatsApp
    ws2 = wb.create_sheet("Solo WhatsApp")
    ws2.append(['#', 'Negocio', 'Telefono', 'WhatsApp', 'Direccion', 'Website'])
    for col in range(1, 7):
        cell = ws2.cell(row=1, column=col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = gris_header
        cell.alignment = Alignment(horizontal='center')

    contador = 1
    for lead in leads:
        if lead.get('whatsapp') in ('SI', 'PROBABLE'):
            ws2.append([
                contador,
                lead.get('nombre', ''),
                lead.get('telefono', ''),
                lead.get('whatsapp', ''),
                lead.get('direccion', ''),
                lead.get('website', ''),
            ])
            contador += 1

    for col, ancho in enumerate([5, 32, 16, 14, 38, 38], 1):
        ws2.column_dimensions[get_column_letter(col)].width = ancho

    wb.save(archivo)
    print(f"\n  Excel guardado: {archivo}")
    print(f"  Total leads: {total}")
    print(f"  Con WhatsApp confirmado: {con_wa}")
    print(f"  WhatsApp probable: {probable}")

def main():
    query = f"{TIPO_NEGOCIO} en {CIUDAD}"
    url   = f"https://www.google.com/maps/search/{query.replace(' ', '+')}"

    print("=" * 60)
    print("  JERCOL TECHNOLOGIES — Buscador de Leads")
    print("=" * 60)
    print(f"  Buscando: {query}")
    print(f"  Objetivo: {MAX_RESULTS} negocios")
    print(f"  Salida:   {ARCHIVO}")
    print("=" * 60)

    leads = []

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,  # False = ves el navegador — cambia a True para modo silencioso
            args=[
                '--no-sandbox',
                '--disable-blink-features=AutomationControlled',
                '--disable-infobars',
            ]
        )
        context = browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
            viewport={'width': 1366, 'height': 768},
            locale='es-CO',
            timezone_id='America/Bogota',
        )

        # Anti-deteccion
        context.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

        page = context.new_page()

        print("\n  Abriendo Google Maps...")
        delay((2.0, 3.0))
        page.goto(url, wait_until='domcontentloaded', timeout=60000)
        delay((3.0, 5.0))

        # Cerrar dialogo de cookies si aparece
        try:
            page.click('button:has-text("Aceptar todo")', timeout=5000)
            delay((1.0, 2.0))
        except Exception:
            pass
        try:
            page.click('button:has-text("Accept all")', timeout=3000)
            delay((1.0, 2.0))
        except Exception:
            pass

        print("  Cargando resultados...\n")

        procesados = set()
        intentos_sin_nuevos = 0

        while len(leads) < MAX_RESULTS and intentos_sin_nuevos < 5:
            # Buscar items en el panel lateral
            try:
                items = page.locator('[role="article"]').all()
            except Exception:
                items = []

            nuevos_en_iteracion = 0

            for item in items:
                if len(leads) >= MAX_RESULTS:
                    break
                try:
                    nombre_preview = item.locator('div[class*="fontHeadlineSmall"]').first.inner_text(timeout=2000).strip()
                    if not nombre_preview or nombre_preview in procesados:
                        continue
                    procesados.add(nombre_preview)
                except Exception:
                    continue

                print(f"  [{len(leads)+1}/{MAX_RESULTS}] Procesando: {nombre_preview[:45]}...")

                try:
                    item.click(timeout=5000)
                    delay(DELAY_ENTRE_CLICKS)

                    datos = extraer_detalle_negocio(page)
                    datos['nombre'] = nombre_preview

                    # Verificar WhatsApp en web si tiene website
                    wa_web = 'sin_web'
                    if datos['website'] != 'N/A':
                        print(f"         Verificando website...")
                        wa_web = buscar_whatsapp_en_web(page, datos['website'])
                        # Volver a Google Maps
                        page.go_back(wait_until='domcontentloaded', timeout=15000)
                        delay((2.0, 3.5))
                        item.click(timeout=5000)
                        delay((1.5, 2.5))

                    datos['whatsapp'] = determinar_whatsapp(datos['telefono'], wa_web)
                    datos['notas'] = f'Web: {wa_web}' if wa_web not in ('sin_web', 'error') else ''

                    leads.append(datos)
                    nuevos_en_iteracion += 1

                    emoji = '✓' if datos['whatsapp'] in ('SI', 'PROBABLE') else '·'
                    print(f"         {emoji} Tel: {datos['telefono']} | WA: {datos['whatsapp']}")

                    # Volver al listado
                    try:
                        page.go_back(wait_until='domcontentloaded', timeout=10000)
                        delay((1.5, 2.5))
                    except Exception:
                        page.goto(url, wait_until='networkidle', timeout=20000)
                        delay((3.0, 4.0))

                except Exception as e:
                    print(f"         Error: {e}")
                    try:
                        page.go_back(wait_until='domcontentloaded', timeout=10000)
                        delay((1.5, 2.5))
                    except Exception:
                        pass
                    continue

            if nuevos_en_iteracion == 0:
                intentos_sin_nuevos += 1
            else:
                intentos_sin_nuevos = 0

            # Scroll para cargar mas resultados
            if len(leads) < MAX_RESULTS:
                print(f"\n  Cargando mas resultados (scroll)...")
                try:
                    panel = page.locator('[role="feed"]').first
                    panel.evaluate('el => el.scrollBy(0, 800)')
                    delay(DELAY_ENTRE_SCROLL)
                except Exception:
                    try:
                        page.keyboard.press('End')
                        delay(DELAY_ENTRE_SCROLL)
                    except Exception:
                        pass

        browser.close()

    print(f"\n  Extraccion completada — {len(leads)} negocios encontrados")
    print("  Generando Excel...")
    exportar_excel(leads, ARCHIVO)
    print("\n  LISTO. Revisa el archivo:", ARCHIVO)
    print("=" * 60)

if __name__ == '__main__':
    main()
